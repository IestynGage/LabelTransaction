from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from models import Transaction


def export_transactions(
    transactions: List[Transaction],
    filename: str = "transactions.xlsx",
):
    """
    Export a list of Transaction objects to an Excel spreadsheet.

    Args:
        transactions: List of Transaction objects
        filename: Output Excel filename
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Header row
    headers = [
        "Date",
        "Description",
        "Value",
        "Label",
    ]

    ws.append(headers)

    # Make headers bold
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Add transaction rows
    for t in transactions:
        ws.append([
            t.date,
            t.desc,
            t.value,
            t.label,
            t.transactionType,
        ])

    # Auto-size columns
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column

        for cell in column_cells:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Save file
    wb.save(filename)

    print(f"Excel file saved as: {filename}")