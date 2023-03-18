from openpyxl import load_workbook

INCOME = "I"
EXPENSE = "E"
TEMPLATE_FILE_PATH = "resources/template.xlsx"

EXPENSE_COLUMN = 4
INCOME_COLUMN = 12


class Transactions:
    def __init__(self):
        self.labels = []
        self.wb = load_workbook(TEMPLATE_FILE_PATH)

    def addTransaction(
        self, date: str, desc: str, value: int, label: str, transactionType: str
    ):
        self.labels.append(Transaction(date, desc, value, label, transactionType))

    def addTransactions(self, transactions):
        self.labels = self.labels + transactions.labels

    def convertToExcel(self):
        ws = self.wb.active
        incomeRow = 3
        expenseRow = 3
        for transaction in self.labels:
            if transaction.transactionType == EXPENSE:
                ws.cell(row=expenseRow, column=EXPENSE_COLUMN).value = transaction.date
                ws.cell(
                    row=expenseRow, column=EXPENSE_COLUMN + 1
                ).value = transaction.desc
                ws.cell(
                    row=expenseRow, column=EXPENSE_COLUMN + 2
                ).value = transaction.value
                ws.cell(
                    row=expenseRow, column=EXPENSE_COLUMN + 3
                ).value = transaction.label
                expenseRow = expenseRow + 1
            else:
                ws.cell(row=incomeRow, column=INCOME_COLUMN).value = transaction.date
                ws.cell(
                    row=incomeRow, column=INCOME_COLUMN + 1
                ).value = transaction.desc
                ws.cell(
                    row=incomeRow, column=INCOME_COLUMN + 2
                ).value = transaction.value
                ws.cell(
                    row=incomeRow, column=INCOME_COLUMN + 3
                ).value = transaction.label
                incomeRow = incomeRow + 1

        self.wb.save("spending.xlsx")
        print("Export spending.xlsx succesfull")


class Transaction:
    def __init__(
        self, date: str, desc: str, value: int, label: str, transactionType: str
    ):
        self.date = date
        self.desc = desc
        self.value = value
        self.label = label
        self.transactionType = transactionType
